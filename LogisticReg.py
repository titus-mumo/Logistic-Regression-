'''The model has an accuracy of about 88.5% '''
# importing modules
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from openpyxl import load_workbook
import pandas as pd
#read the workbook
wb = load_workbook('data.xlsx')
print(wb.sheetnames)
#extract the appropriate sheet
data = wb['STEM Data']
#extract the appropriate columns and rows
rows_list = []
for row in data:
    cols = []
    for col in row[0:6]:
        cols.append(col.value)
    rows_list.append(cols)
#using pandas dataframe for data holding
rawdata = pd.DataFrame(data = rows_list[1: ], index = None, columns = rows_list[0])
#converting pandas dataframe to numpy array
arr = rawdata.to_numpy()
#determine the shape of the numpy array
print(arr.shape)
#extracting the target of the dataset
result = arr[ :, 0]
print(result)
#extracting the features of the dataset
data = arr[ :, 1:]
print(data)
#spitting data to training and testing datasets
x_train, x_test, y_train, y_test = train_test_split(data, result, test_size=0.23, random_state=2)
#deploying logistic regression ML model
clf = LogisticRegression()
clf.fit(x_train, y_train)
#using the model to predict target values
print(clf.predict(x_test))
predictions = clf.predict(x_test)
for i, prediction in enumerate(predictions):   
    print('Predicted: %s, Target: %s' % (prediction, y_test[i]))

#checking accuracy using test data
y_pred = clf.predict(x_test)
print(f"Accuracy is: {(accuracy_score(y_test, y_pred))*100}")
#checking accuracy using train data
y_pred2 = clf.predict(x_train)
print(f"Accuracy is: {(accuracy_score(y_train, y_pred2))*100}")

